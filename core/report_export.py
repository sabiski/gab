"""Export de rapports pharmacie (Excel, PDF, CSV)."""

from __future__ import annotations

import csv
import io
from datetime import date, datetime

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


# Couleurs Gab'Pharma
GREEN = "228545"
GREEN_DARK = "015533"
LIGHT_GREEN = "E8F5E9"
HEADER_FILL = PatternFill("solid", fgColor=GREEN)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color=GREEN_DARK)
SECTION_FONT = Font(bold=True, size=11, color=GREEN_DARK)
MONEY_FMT = '#,##0 " F"'
INT_FMT = "#,##0"
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fmt_money(value) -> str:
    if value is None:
        return "—"
    try:
        return f"{int(value):,}".replace(",", " ") + " F"
    except (TypeError, ValueError):
        return str(value)


def _slug_filename(pharmacy, prefix: str, ext: str) -> str:
    slug = pharmacy.slug if pharmacy else "pharma"
    stamp = timezone.localdate().strftime("%Y%m%d")
    return f"{prefix}-{slug}-{stamp}.{ext}"


def _style_header_row(ws, row: int, ncol: int):
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws, min_w=10, max_w=42):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def build_pharmacy_stats_workbook(payload: dict) -> Workbook:
    """Génère un classeur Excel formaté pour le rapport statistiques."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    pharmacy_name = payload.get("pharmacy_name") or "Pharmacie"
    days = payload.get("period_days", 30)
    start: date = payload.get("start_date")
    end: date = payload.get("end_date")
    kpis = payload.get("kpis", {})
    generated = payload.get("generated_at") or timezone.now()

    # --- En-tête ---
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"Gab'Pharma — Rapport statistiques"
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:D2")
    ws["A2"].value = pharmacy_name
    ws["A2"].font = Font(bold=True, size=12)

    period_label = f"Période : {days} jours"
    if start and end:
        period_label += f" ({start.strftime('%d/%m/%Y')} → {end.strftime('%d/%m/%Y')})"
    ws.merge_cells("A3:D3")
    ws["A3"].value = period_label
    ws["A3"].font = Font(color="666666", size=10)

    ws.merge_cells("A4:D4")
    ws["A4"].value = f"Généré le {timezone.localtime(generated).strftime('%d/%m/%Y à %H:%M')}"
    ws["A4"].font = Font(color="999999", size=9, italic=True)

    row = 6
    ws.cell(row=row, column=1, value="Indicateurs clés").font = SECTION_FONT
    row += 1

    kpi_rows = [
        ("CA période", kpis.get("ca_period"), MONEY_FMT),
        ("Commandes période", kpis.get("orders_period"), INT_FMT),
        ("Panier moyen", kpis.get("avg_basket"), MONEY_FMT),
        ("Marge estimée", kpis.get("margin"), MONEY_FMT),
        ("Marge (%)", kpis.get("margin_pct"), '0"%"'),
        ("Clients servis", kpis.get("clients_served"), INT_FMT),
        ("Livraisons effectuées", kpis.get("deliveries_done"), INT_FMT),
        ("CA total (historique)", kpis.get("ca_total"), MONEY_FMT),
        ("Commandes totales", kpis.get("orders_total"), INT_FMT),
        ("Valeur stock", kpis.get("stock_value"), MONEY_FMT),
    ]
    deliv = payload.get("deliv") or {}
    avg_min = deliv.get("avg_delivery_min")
    on_time = deliv.get("on_time_rate")
    kpi_rows.extend(
        [
            ("Délai moyen livraison (min)", avg_min if avg_min is not None else "—", None),
            ("Taux livraison à temps (%)", on_time if on_time is not None else "—", '0"%"'),
        ]
    )

    ws.cell(row=row, column=1, value="Indicateur")
    ws.cell(row=row, column=2, value="Valeur")
    _style_header_row(ws, row, 2)
    row += 1
    for label, val, fmt in kpi_rows:
        ws.cell(row=row, column=1, value=label).border = BORDER
        cell = ws.cell(row=row, column=2, value=val if val != "—" else "—")
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right")
        if fmt and val != "—":
            cell.number_format = fmt
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Évolution quotidienne").font = SECTION_FONT
    row += 1
    labels = payload.get("daily_labels") or []
    orders_data = payload.get("orders_data") or []
    ca_data = payload.get("ca_data") or []
    ws.cell(row=row, column=1, value="Date")
    ws.cell(row=row, column=2, value="Commandes")
    ws.cell(row=row, column=3, value="CA (F)")
    _style_header_row(ws, row, 3)
    row += 1
    for i, lbl in enumerate(labels):
        ws.cell(row=row, column=1, value=lbl).border = BORDER
        c_ord = ws.cell(row=row, column=2, value=orders_data[i] if i < len(orders_data) else 0)
        c_ord.border = BORDER
        c_ord.number_format = INT_FMT
        c_ca = ws.cell(row=row, column=3, value=ca_data[i] if i < len(ca_data) else 0)
        c_ca.border = BORDER
        c_ca.number_format = MONEY_FMT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Modes de paiement").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Mode")
    ws.cell(row=row, column=2, value="Montant (F)")
    ws.cell(row=row, column=3, value="Transactions")
    _style_header_row(ws, row, 3)
    row += 1
    for p in payload.get("pay_rows") or []:
        ws.cell(row=row, column=1, value=p.get("label", "")).border = BORDER
        c1 = ws.cell(row=row, column=2, value=p.get("total", 0))
        c1.border = BORDER
        c1.number_format = MONEY_FMT
        c2 = ws.cell(row=row, column=3, value=p.get("count", 0))
        c2.border = BORDER
        c2.number_format = INT_FMT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Ventes par catégorie").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Catégorie")
    ws.cell(row=row, column=2, value="CA (F)")
    _style_header_row(ws, row, 2)
    row += 1
    for c in payload.get("cat_rows") or []:
        ws.cell(row=row, column=1, value=c.get("category", "")).border = BORDER
        cell = ws.cell(row=row, column=2, value=c.get("total", 0))
        cell.border = BORDER
        cell.number_format = MONEY_FMT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Commandes par statut").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Statut")
    ws.cell(row=row, column=2, value="Nombre")
    _style_header_row(ws, row, 2)
    row += 1
    for s in payload.get("status_rows") or []:
        ws.cell(row=row, column=1, value=s.get("label", "")).border = BORDER
        cell = ws.cell(row=row, column=2, value=s.get("count", 0))
        cell.border = BORDER
        cell.number_format = INT_FMT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top produits vendus").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Produit")
    ws.cell(row=row, column=2, value="Quantité")
    _style_header_row(ws, row, 2)
    row += 1
    for m in payload.get("top_meds") or []:
        ws.cell(row=row, column=1, value=m.get("medicine_name", "")).border = BORDER
        cell = ws.cell(row=row, column=2, value=m.get("qty", 0))
        cell.border = BORDER
        cell.number_format = INT_FMT
        row += 1

    _auto_width(ws)
    ws.freeze_panes = "A6"
    return wb


def pharmacy_stats_xlsx_response(payload: dict, pharmacy) -> HttpResponse:
    wb = build_pharmacy_stats_workbook(payload)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_slug_filename(pharmacy, "rapport-stats", "xlsx")}"'
    return response


def pharmacy_stats_pdf_response(payload: dict, pharmacy) -> HttpResponse:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title="Rapport statistiques",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleGP",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor(f"#{GREEN_DARK}"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "SubGP",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=2,
    )
    section_style = ParagraphStyle(
        "SecGP",
        parent=styles["Heading2"],
        fontSize=11,
        textColor=colors.HexColor(f"#{GREEN}"),
        spaceBefore=10,
        spaceAfter=4,
    )

    story = []
    pharmacy_name = payload.get("pharmacy_name") or "Pharmacie"
    days = payload.get("period_days", 30)
    start: date = payload.get("start_date")
    end: date = payload.get("end_date")
    generated = payload.get("generated_at") or timezone.now()
    kpis = payload.get("kpis", {})

    story.append(Paragraph("Gab'Pharma — Rapport statistiques", title_style))
    story.append(Paragraph(f"<b>{pharmacy_name}</b>", styles["Normal"]))
    period_txt = f"Période : {days} jours"
    if start and end:
        period_txt += f" ({start.strftime('%d/%m/%Y')} → {end.strftime('%d/%m/%Y')})"
    story.append(Paragraph(period_txt, subtitle_style))
    story.append(
        Paragraph(
            f"Généré le {timezone.localtime(generated).strftime('%d/%m/%Y à %H:%M')}",
            subtitle_style,
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    def _table(data, col_widths=None):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{GREEN}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{LIGHT_GREEN}")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return t

    story.append(Paragraph("Indicateurs clés", section_style))
    deliv = payload.get("deliv") or {}
    kpi_data = [
        ["Indicateur", "Valeur"],
        ["CA période", _fmt_money(kpis.get("ca_period"))],
        ["Commandes période", str(kpis.get("orders_period", 0))],
        ["Panier moyen", _fmt_money(kpis.get("avg_basket"))],
        ["Marge estimée", _fmt_money(kpis.get("margin"))],
        ["Marge (%)", f"{kpis.get('margin_pct', 0)} %"],
        ["Clients servis", str(kpis.get("clients_served", 0))],
        ["Livraisons effectuées", str(kpis.get("deliveries_done", 0))],
        ["CA total", _fmt_money(kpis.get("ca_total"))],
        [
            "Délai moyen livraison",
            f"{deliv.get('avg_delivery_min')} min" if deliv.get("avg_delivery_min") else "—",
        ],
        [
            "Taux à temps",
            f"{deliv.get('on_time_rate', 0)} %" if deliv.get("on_time_rate") is not None else "—",
        ],
    ]
    story.append(_table(kpi_data, [10 * cm, 7 * cm]))

    labels = payload.get("daily_labels") or []
    orders_data = payload.get("orders_data") or []
    ca_data = payload.get("ca_data") or []
    if labels:
        story.append(Paragraph("Évolution quotidienne", section_style))
        daily = [["Date", "Commandes", "CA"]]
        for i, lbl in enumerate(labels):
            daily.append(
                [
                    lbl,
                    str(orders_data[i] if i < len(orders_data) else 0),
                    _fmt_money(ca_data[i] if i < len(ca_data) else 0),
                ]
            )
        story.append(_table(daily, [4 * cm, 4 * cm, 9 * cm]))

    pay_rows = payload.get("pay_rows") or []
    if pay_rows:
        story.append(Paragraph("Modes de paiement", section_style))
        pay_data = [["Mode", "Montant", "Transactions"]]
        for p in pay_rows:
            pay_data.append([p.get("label", ""), _fmt_money(p.get("total")), str(p.get("count", 0))])
        story.append(_table(pay_data, [7 * cm, 5 * cm, 5 * cm]))

    cat_rows = payload.get("cat_rows") or []
    if cat_rows:
        story.append(Paragraph("Ventes par catégorie", section_style))
        cat_data = [["Catégorie", "CA"]]
        for c in cat_rows:
            cat_data.append([c.get("category", ""), _fmt_money(c.get("total"))])
        story.append(_table(cat_data, [10 * cm, 7 * cm]))

    status_rows = payload.get("status_rows") or []
    if status_rows:
        story.append(Paragraph("Commandes par statut", section_style))
        st_data = [["Statut", "Nombre"]]
        for s in status_rows:
            st_data.append([s.get("label", ""), str(s.get("count", 0))])
        story.append(_table(st_data, [10 * cm, 7 * cm]))

    top_meds = payload.get("top_meds") or []
    if top_meds:
        story.append(Paragraph("Top produits vendus", section_style))
        top_data = [["Produit", "Quantité"]]
        for m in top_meds:
            top_data.append([m.get("medicine_name", ""), str(m.get("qty", 0))])
        story.append(_table(top_data, [12 * cm, 5 * cm]))

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_slug_filename(pharmacy, "rapport-stats", "pdf")}"'
    return response


def pharmacy_stats_csv_response(payload: dict, pharmacy) -> HttpResponse:
    """CSV UTF-8 avec BOM et séparateur ; pour Excel français."""
    buf = io.StringIO()
    buf.write("\ufeff")
    writer = csv.writer(buf, delimiter=";")
    kpis = payload.get("kpis", {})
    days = payload.get("period_days", 30)
    pharmacy_name = payload.get("pharmacy_name") or "Pharmacie"

    writer.writerow(["Gab'Pharma — Rapport statistiques"])
    writer.writerow(["Pharmacie", pharmacy_name])
    writer.writerow(["Période (jours)", days])
    writer.writerow([])
    writer.writerow(["Indicateur", "Valeur"])
    writer.writerow(["CA période (F)", kpis.get("ca_period", 0)])
    writer.writerow(["Commandes période", kpis.get("orders_period", 0)])
    writer.writerow(["Panier moyen (F)", kpis.get("avg_basket", 0)])
    writer.writerow(["Marge estimée (F)", kpis.get("margin", 0)])
    writer.writerow(["Marge (%)", kpis.get("margin_pct", 0)])
    writer.writerow(["Clients servis", kpis.get("clients_served", 0)])
    writer.writerow(["Livraisons effectuées", kpis.get("deliveries_done", 0)])
    writer.writerow(["CA total (F)", kpis.get("ca_total", 0)])
    writer.writerow([])
    writer.writerow(["Date", "Commandes", "CA (F)"])
    labels = payload.get("daily_labels") or []
    orders_data = payload.get("orders_data") or []
    ca_data = payload.get("ca_data") or []
    for i, lbl in enumerate(labels):
        writer.writerow(
            [
                lbl,
                orders_data[i] if i < len(orders_data) else 0,
                ca_data[i] if i < len(ca_data) else 0,
            ]
        )
    writer.writerow([])
    writer.writerow(["Mode paiement", "Montant F", "Transactions"])
    for p in payload.get("pay_rows") or []:
        writer.writerow([p.get("label"), p.get("total"), p.get("count")])
    writer.writerow([])
    writer.writerow(["Catégorie", "CA (F)"])
    for c in payload.get("cat_rows") or []:
        writer.writerow([c.get("category"), c.get("total")])
    writer.writerow([])
    writer.writerow(["Statut", "Nombre"])
    for s in payload.get("status_rows") or []:
        writer.writerow([s.get("label"), s.get("count")])
    writer.writerow([])
    writer.writerow(["Produit", "Quantité"])
    for m in payload.get("top_meds") or []:
        writer.writerow([m.get("medicine_name"), m.get("qty")])

    response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{_slug_filename(pharmacy, "rapport-stats", "csv")}"'
    return response
